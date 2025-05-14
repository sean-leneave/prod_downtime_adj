import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime

# Create production forecast template
def create_production_template():
    # Read the CSV template
    df = pd.read_csv('template_production_forecast.csv')
    
    # Create a new Excel file
    wb = Workbook()
    ws = wb.active
    ws.title = "Production Forecast"
    
    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    instructions_font = Font(bold=True, size=11)
    
    note_font = Font(italic=True, size=10)
    
    # Add instructions at the top
    ws.merge_cells('A1:E1')
    ws['A1'] = "Production Forecast Template"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:E2')
    ws['A2'] = "Enter monthly production forecasts for each well below"
    ws['A2'].font = instructions_font
    
    # Add header row with better descriptions
    headers = [
        "Well Name", 
        "Date (MM/DD/YY)", 
        "Oil Rate (bopd)", 
        "Gas Rate (mcfd)", 
        "Water Rate (bwpd)"
    ]
    
    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Add sample data
    data = [
        ["WellA", "1/31/2024", 100, 500, 200],
        ["WellA", "2/29/2024", 95, 480, 210],
        ["WellA", "3/31/2024", 90, 470, 220],
        ["WellB", "1/31/2024", 100, 500, 200],
        ["WellB", "2/29/2024", 95, 480, 210],
        ["WellB", "3/31/2024", 90, 470, 220],
    ]
    
    for row_idx, row_data in enumerate(data, 5):
        for col_idx, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = cell_value
            # Format date cells
            if col_idx == 2:
                cell.number_format = 'mm/dd/yyyy'
    
    # Add more empty rows
    for row_idx in range(len(data) + 5, len(data) + 15):
        for col_idx in range(1, 6):
            ws.cell(row=row_idx, column=col_idx).value = ""
    
    # Add notes
    note_row = len(data) + 16
    ws.merge_cells(f'A{note_row}:E{note_row}')
    ws[f'A{note_row}'] = "Notes: Dates should be in MM/DD/YYYY format. Rates are daily averages for the month."
    ws[f'A{note_row}'].font = note_font
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    
    # Save the file
    wb.save('template_production_forecast.xlsx')
    print("Production forecast template created as template_production_forecast.xlsx")

# Create downtime forecast template
def create_downtime_template():
    # Read the CSV template
    df = pd.read_csv('template_downtime_forecast.csv')
    
    # Create a new Excel file
    wb = Workbook()
    ws = wb.active
    ws.title = "Downtime Forecast"
    
    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    instructions_font = Font(bold=True, size=11)
    
    note_font = Font(italic=True, size=10)
    
    # Add instructions at the top
    ws.merge_cells('A1:C1')
    ws['A1'] = "Downtime Forecast Template"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:C2')
    ws['A2'] = "Enter monthly downtime percentage forecasts below"
    ws['A2'].font = instructions_font
    
    # Add header row with better descriptions
    headers = [
        "Scenario", 
        "Date (MM/DD/YY)", 
        "Downtime Percentage (0-100)"
    ]
    
    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Add sample data
    data = [
        ["OL3", "1/31/2024", 5],
        ["OL3", "2/29/2024", 7],
        ["OL3", "3/31/2024", 10],
    ]
    
    for row_idx, row_data in enumerate(data, 5):
        for col_idx, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = cell_value
            # Format date cells
            if col_idx == 2:
                cell.number_format = 'mm/dd/yyyy'
            # Format percentage cells
            if col_idx == 3:
                cell.number_format = '0.00'
    
    # Add more empty rows
    for row_idx in range(len(data) + 5, len(data) + 15):
        for col_idx in range(1, 4):
            ws.cell(row=row_idx, column=col_idx).value = ""
    
    # Add notes
    note_row = len(data) + 16
    ws.merge_cells(f'A{note_row}:C{note_row}')
    ws[f'A{note_row}'] = "Notes: Dates should be in MM/DD/YYYY format. Downtime % should be between 0-100."
    ws[f'A{note_row}'].font = note_font
    
    note_row += 1
    ws.merge_cells(f'A{note_row}:C{note_row}')
    ws[f'A{note_row}'] = "Example: 5 means 5% downtime (not running 5% of the time)"
    ws[f'A{note_row}'].font = note_font
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 25
    
    # Save the file
    wb.save('template_downtime_forecast.xlsx')
    print("Downtime forecast template created as template_downtime_forecast.xlsx")

if __name__ == "__main__":
    create_production_template()
    create_downtime_template()
    print("Template conversion complete!") 