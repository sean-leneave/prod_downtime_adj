import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import io
import zipfile
from utils import (
    standardize_downtime_percentage,
    interpolate_values,
    add_one_month,
    standardize_to_end_of_month
)
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker

# Constants
DEBUG_MODE = False  # Set to True to show debugging info messages

COLORS = {
    'background': '#F8F9FA',
    'text': '#2C3E50',
    'oil': '#2ECC71',
    'gas': '#E74C3C',
    'water': '#3498DB',
    'liquid': '#34495E'
}

PLOT_LAYOUT = {
    'font': {
        'family': 'Arial, sans-serif',
        'size': 12,
        'color': COLORS['text']
    },
    'plot_bgcolor': COLORS['background'],
    'hovermode': 'x unified'
}

# Set page config
st.set_page_config(
    page_title="Production Downtime Adjustment",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Constants
MAX_YEARS = 30
MAX_MONTHS = MAX_YEARS * 12

# Initialize session state
if 'forecast_data' not in st.session_state:
    st.session_state.forecast_data = ""
if 'downtime_data' not in st.session_state:
    st.session_state.downtime_data = ""
if 'processed_data' not in st.session_state:
    st.session_state.processed_data = None
if 'figure' not in st.session_state:
    st.session_state.figure = None
if 'batch_data' not in st.session_state:
    st.session_state.batch_data = None
if 'batch_wells' not in st.session_state:
    st.session_state.batch_wells = []
if 'selected_well' not in st.session_state:
    st.session_state.selected_well = None
if 'well_forecast_data' not in st.session_state:
    st.session_state.well_forecast_data = {}
if 'processed_wells' not in st.session_state:
    st.session_state.processed_wells = {}
if 'mosaic_zip_buffer' not in st.session_state:
    st.session_state.mosaic_zip_buffer = None
if 'mosaic_zip_filename' not in st.session_state:
    st.session_state.mosaic_zip_filename = None
if 'generated_export_bytes' not in st.session_state:
    st.session_state.generated_export_bytes = None
if 'generated_export_filename' not in st.session_state:
    st.session_state.generated_export_filename = None


def process_and_store_data():
    """Process input data and store results in session state"""
    if not st.session_state.forecast_data or not st.session_state.downtime_data:
        st.error("Please provide both forecast and downtime data.")
        return False

    try:
        # Load dataframes
        try:
            forecast_df = pd.read_csv(
                io.StringIO(
                    st.session_state.forecast_data),
                sep='\t')
            if len(forecast_df.columns) == 1:
                # Try comma separator if tab didn't work
                forecast_df = pd.read_csv(
                    io.StringIO(
                        st.session_state.forecast_data),
                    sep=',')
                if DEBUG_MODE:
                    st.info("Detected comma-separated values for forecast data.")
        except Exception as e:
            st.error(f"Error parsing forecast data: {str(e)}")
            return False

        try:
            downtime_df = pd.read_csv(
                io.StringIO(
                    st.session_state.downtime_data),
                sep='\t')
            if len(downtime_df.columns) == 1:
                downtime_df = pd.read_csv(
                    io.StringIO(
                        st.session_state.downtime_data),
                    sep=',')
                if DEBUG_MODE:
                    st.info("Detected comma-separated values for downtime data.")
        except Exception as e:
            st.error(f"Error parsing downtime data: {str(e)}")
            return False

        # Print loaded columns for debugging
        if DEBUG_MODE:
            st.info(
                f"Loaded forecast columns: {', '.join(forecast_df.columns)}")

        # Handle column mapping for forecast data
        standard_columns = {
            'Well Name': [
                'well_name', 'well name', 'wellname', 'well', 'name', 'entity', 'Well Name'
            ], 'date': ['date', 'Date'], 'oil_rate': [
                'oil_rate', 'oil rate', 'oilrate', 'oil', 'Oil Rate'], 'gas_rate': [
                'gas_rate', 'gas rate', 'gasrate', 'gas', 'Gas Rate'], 'water_rate': [
                'water_rate', 'water rate', 'waterrate', 'wtrrate', 'water', 'wtr', 'Water Rate']}

        orig_columns = list(forecast_df.columns)
        normalized_columns = [str(col).strip().lower()
                              for col in forecast_df.columns]

        # Map columns to standard names
        col_map = {}
        for std_col, patterns in standard_columns.items():
            for i, col_lower in enumerate(normalized_columns):
                orig_col = orig_columns[i]

                # check for exact match with any pattern
                if col_lower in [p.lower() for p in patterns]:
                    col_map[orig_col] = std_col
                    break

                # check for partial matches (skip if already mapped)
                if orig_col not in col_map:
                    for pattern in patterns:
                        if pattern.lower() in col_lower:
                            col_map[orig_col] = std_col
                            break

                # Break out of the loop if mapped this column
                if orig_col in col_map:
                    break

        # Report mapping for debugging
        if col_map and DEBUG_MODE:
            mapping_info = ", ".join(
                [f"{orig} → {new}" for orig, new in col_map.items()])
            st.info(f"Column mapping: {mapping_info}")

        # Apply column mapping
        forecast_df = forecast_df.rename(columns=col_map)

        # Check for required columns
        forecast_columns = [
            'Well Name',
            'date',
            'oil_rate',
            'gas_rate',
            'water_rate']
        missing_cols = [
            col for col in forecast_columns if col not in forecast_df.columns]

        # Try alternate mapping if columns are still missing
        if 'Well Name' in missing_cols and 'well_name' in forecast_df.columns:
            forecast_df['Well Name'] = forecast_df['well_name']
            missing_cols.remove('Well Name')

        if missing_cols:
            st.error(
                f"Forecast file is missing required columns: {', '.join(missing_cols)}."
                "Please check your file.")
            if DEBUG_MODE:
                st.info(f"Available columns: {', '.join(forecast_df.columns)}")
            return False

        # Handle similar mapping for downtime data if needed
        downtime_columns = ['Scenario', 'date', 'downtime_pct']
        for req in downtime_columns:
            if req not in downtime_df.columns:
                downtime_df[req] = '' if req == 'Scenario' or req == 'date' else None
        downtime_df = downtime_df[downtime_columns]
        downtime_df['Scenario'] = downtime_df['Scenario'].astype(str)

        # Check if DataFrames are empty
        if forecast_df.empty or downtime_df.empty:
            st.error("Forecast or downtime data is empty or invalid.")
            return False

        # Process each unique well independently
        st.session_state.processed_wells = {}
        for well in forecast_df['Well Name'].unique():
            well_df = forecast_df[forecast_df['Well Name'] == well].copy()
            # Defensive check for downtime scenario
            scenario = ''
            if not downtime_df.empty \
                    and 'Scenario' in downtime_df.columns \
                    and downtime_df['Scenario'].notnull(
                    ).any():
                scenario = downtime_df['Scenario'].iloc[0]
            dt_df = downtime_df[downtime_df['Scenario'] == scenario].copy(
            ) if scenario else downtime_df.copy()
            # Use process_input_data to get processed DataFrames
            well_forecast_str = well_df.to_csv(sep='\t', index=False)
            dt_str = dt_df.to_csv(sep='\t', index=False)
            f_df, d_df = process_input_data(well_forecast_str, dt_str)
            if f_df is None or d_df is None or f_df.empty or d_df.empty:
                continue
            errors, warnings = validate_input_data(f_df, d_df)
            if errors:
                continue
            df = f_df.join(d_df, how='left')
            df = df.sort_index()
            df_out, df_interpolate = process_production_data(df)

            # Add defensive check for empty result DataFrames
            if df_out is None or df_out.empty or len(df_out.index) == 0:
                # Skip this well and move to the next one
                continue

            st.session_state.processed_wells[well] = {
                'data': df_out,
                'figure': create_plotly_figure(df_out, well_name=well)
            }
        # Set the first well as selected and update processed_data/figure
        if st.session_state.processed_wells:
            first_well = next(iter(st.session_state.processed_wells))
            st.session_state.selected_well = first_well
            st.session_state.processed_data = st.session_state.processed_wells[first_well]['data']
            st.session_state.figure = st.session_state.processed_wells[first_well]['figure']
        else:
            st.session_state.processed_data = None
            st.session_state.figure = None
        return True
    except Exception as e:
        st.error(f"Error processing data: {str(e)}")
        return False


def process_input_data(forecast_data, downtime_data):
    """Process input data into proper format for calculations."""
    if not forecast_data or not downtime_data:
        st.error("Please provide both forecast and downtime data.")
        return None, None

    try:
        # Convert strings to dataframes
        try:
            forecast_df = pd.read_csv(io.StringIO(forecast_data), sep='\t')
            if len(forecast_df.columns) == 1:
                forecast_df = pd.read_csv(io.StringIO(forecast_data), sep=',')
                if DEBUG_MODE:
                    st.info(
                        "Detected comma-separated values for forecast data instead of tabs."
                        "Processing anyway.")
        except Exception as e:
            st.error(
                f"Error parsing forecast data: {str(e)}"
                "Please ensure data is tab-separated with proper headers.")
            return None, None

        try:
            downtime_df = pd.read_csv(io.StringIO(downtime_data), sep='\t')
            if len(downtime_df.columns) == 1:
                downtime_df = pd.read_csv(io.StringIO(downtime_data), sep=',')
                if DEBUG_MODE:
                    st.info(
                        "Detected comma-separated values for downtime data instead of tabs."
                        "Processing anyway.")
        except Exception as e:
            st.error(
                f"Error parsing downtime data: {str(e)}"
                "Please ensure data is tab-separated with proper headers.")
            return None, None

        # Print loaded columns for debugging
        if DEBUG_MODE:
            st.info(f"Loaded columns: {', '.join(forecast_df.columns)}")

        # Handle column mapping for forecast data
        standard_columns = {
            'Well Name': [
                'well_name', 'well name', 'wellname', 'well', 'name', 'entity', 'Well Name'], 'date': [
                'date', 'Date'], 'oil_rate': [
                'oil_rate', 'oil rate', 'oilrate', 'oil', 'Oil Rate'], 'gas_rate': [
                    'gas_rate', 'gas rate', 'gasrate', 'gas', 'Gas Rate'], 'water_rate': [
                'water_rate', 'water rate', 'waterrate', 'wtrrate', 'water', 'wtr', 'Water Rate']}

        orig_columns = list(forecast_df.columns)
        normalized_columns = [str(col).strip().lower()
                              for col in forecast_df.columns]

        # Map columns to standard names
        col_map = {}
        for std_col, patterns in standard_columns.items():
            for i, col_lower in enumerate(normalized_columns):
                orig_col = orig_columns[i]

                # Check for exact match with any pattern
                if col_lower in [p.lower() for p in patterns]:
                    col_map[orig_col] = std_col
                    break

                # Check for partial matches (skip if already mapped)
                if orig_col not in col_map:
                    for pattern in patterns:
                        if pattern.lower() in col_lower:
                            col_map[orig_col] = std_col
                            break

                # Break out of the loop if mapped this column
                if orig_col in col_map:
                    break

        # Report mapping for debugging
        if col_map and DEBUG_MODE:
            mapping_info = ", ".join(
                [f"{orig} → {new}" for orig, new in col_map.items()])
            st.info(f"Column mapping: {mapping_info}")

        # Apply column mapping
        forecast_df = forecast_df.rename(columns=col_map)

        # Check for required columns
        required_cols = [
            'Well Name',
            'date',
            'oil_rate',
            'gas_rate',
            'water_rate']
        missing_cols = [
            col for col in required_cols if col not in forecast_df.columns]

        # Try alternate mapping if columns are still missing
        if 'Well Name' in missing_cols and 'well_name' in forecast_df.columns:
            forecast_df['Well Name'] = forecast_df['well_name']
            missing_cols.remove('Well Name')

        if missing_cols:
            st.error(
                f"Forecast file is missing required columns: {', '.join(missing_cols)}."
                "Please check your file.")
            if DEBUG_MODE:
                st.info(f"Available columns: {', '.join(forecast_df.columns)}")
            return None, None

        # Check for required columns
        if 'date' not in forecast_df.columns:
            st.error("Forecast data must include a 'date' column.")
            return None, None
        if 'date' not in downtime_df.columns:
            st.error("Downtime data must include a 'date' column.")
            return None, None

        # Convert dates to datetime
        forecast_df['date'] = pd.to_datetime(forecast_df['date'])
        downtime_df['date'] = pd.to_datetime(downtime_df['date'])

        # Standardize dates to end of month
        forecast_df['date'] = forecast_df['date'].apply(
            standardize_to_end_of_month)
        downtime_df['date'] = downtime_df['date'].apply(
            standardize_to_end_of_month)

        # Standardize downtime percentages to decimal format
        if 'downtime_pct' in downtime_df.columns:
            try:
                downtime_df['downtime_pct'] = downtime_df['downtime_pct'].apply(
                    standardize_downtime_percentage)
            except ValueError as e:
                st.error(f"Error in downtime data: {str(e)}")
                return None, None

        # Set index to date
        forecast_df.set_index('date', inplace=True)
        downtime_df.set_index('date', inplace=True)

        # Sort by date
        forecast_df.sort_index(inplace=True)
        downtime_df.sort_index(inplace=True)

        # Fill any NaN values with zeros in numeric columns
        for col in ['oil_rate', 'gas_rate', 'water_rate']:
            if col in forecast_df.columns:
                forecast_df[col] = forecast_df[col].fillna(0)

        if 'downtime_pct' in downtime_df.columns:
            downtime_df['downtime_pct'] = downtime_df['downtime_pct'].fillna(0)

        return forecast_df, downtime_df

    except Exception as e:
        st.error(f"Error processing data: {str(e)}")
        return None, None


def process_batch_data(batch_data, downtime_data):
    """Process batch forecast data with multiple wells."""
    if not batch_data or not downtime_data:
        st.error("Please provide both batch forecast and downtime data.")
        return None, None, None

    try:
        # Convert strings to dataframes
        try:
            batch_df = pd.read_csv(io.StringIO(batch_data), sep='\t')
            if len(batch_df.columns) == 1:
                batch_df = pd.read_csv(io.StringIO(batch_data), sep=',')
                if DEBUG_MODE:
                    st.info("Detected comma-separated values for batch data.")
        except Exception as e:
            st.error(
                f"Error parsing batch data: {str(e)}"
                "Please ensure data is tab-separated with proper headers.")
            return None, None, None

        # Print debug info about loaded columns
        if DEBUG_MODE:
            st.info(
                f"Loaded batch data columns: {', '.join(batch_df.columns)}")

        if batch_df is not None and not batch_df.empty:
            # Create a mapping of original columns to standard names
            standard_columns = {
                'well_name': [
                    'well_name', 'well name', 'wellname', 'well', 'name', 'entity', 'Well Name'], 'date': [
                    'date', 'Date'], 'oil_rate': [
                    'oil_rate', 'oil rate', 'oilrate', 'oil', 'Oil Rate'], 'gas_rate': [
                    'gas_rate', 'gas rate', 'gasrate', 'gas', 'Gas Rate'], 'water_rate': [
                    'water_rate', 'water rate', 'waterrate', 'wtrrate', 'water', 'wtr', 'Water Rate']}

            # Preserve original column names for reporting
            orig_columns = list(batch_df.columns)

            # Create a normalized version of columns for matching (strip,
            # lowercase)
            normalized_columns = [str(col).strip().lower()
                                  for col in batch_df.columns]

            # Map columns to standard names
            col_map = {}
            for std_col, patterns in standard_columns.items():
                for i, col_lower in enumerate(normalized_columns):
                    orig_col = orig_columns[i]

                    # Check for exact match with any pattern
                    if col_lower in [p.lower() for p in patterns]:
                        col_map[orig_col] = std_col
                        break

                    # Check for partial matches
                    for pattern in patterns:
                        if pattern.lower() in col_lower:
                            col_map[orig_col] = std_col
                            break

                    # Break out of the loop if mapped this column
                    if orig_col in col_map:
                        break

            # Report mapping for debugging
            if col_map and DEBUG_MODE:
                mapping_info = ", ".join(
                    [f"{orig} → {new}" for orig, new in col_map.items()])
                st.info(f"Column mapping: {mapping_info}")

            # Apply column mapping
            batch_df = batch_df.rename(columns=col_map)

        required_cols = [
            'well_name',
            'date',
            'oil_rate',
            'gas_rate',
            'water_rate']
        missing_cols = [
            col for col in required_cols if col not in batch_df.columns]

        if missing_cols:
            # Try again with 'Well Name' instead of 'well_name' if that's
            # what's missing
            if 'well_name' in missing_cols and 'Well Name' in batch_df.columns:
                batch_df['well_name'] = batch_df['Well Name']
                missing_cols.remove('well_name')

            if missing_cols:
                st.error(
                    f"Batch forecast file is missing required columns: {', '.join(missing_cols)}."
                    "Please check your file.")
                if DEBUG_MODE:
                    st.info(
                        f"Available columns: {', '.join(batch_df.columns)}")
                return None, None, None

        if batch_df is None or batch_df.empty:
            st.error("Batch forecast file is empty. Please check your file.")
            return None, None, None

        try:
            downtime_df = pd.read_csv(io.StringIO(downtime_data), sep='\t')
            if len(downtime_df.columns) == 1:
                downtime_df = pd.read_csv(io.StringIO(downtime_data), sep=',')
                if DEBUG_MODE:
                    st.info("Detected comma-separated values for downtime data.")
        except Exception as e:
            st.error(
                f"Error parsing downtime data: {str(e)}"
                "Please ensure data is tab-separated with proper headers.")
            return None, None, None

        # Handle the downtime data columns using the same approach
        if downtime_df is not None and not downtime_df.empty:
            # Create a mapping of original columns to standard names for
            # downtime
            downtime_columns = {
                'Scenario': [
                    'scenario',
                    'scenario_name',
                    'case',
                    'name',
                    'downtime_scenario',
                    'Scenario'],
                'date': [
                    'date',
                    'Date'],
                'downtime_pct': [
                    'downtime_pct',
                    'downtime',
                    'dt_pct',
                    'dt',
                    'pct',
                    'percent',
                    'downtime_percent']}

            # Preserve original column names for reporting
            orig_dt_columns = list(downtime_df.columns)

            # Create a normalized version of columns for matching
            normalized_dt_columns = [
                str(col).strip().lower() for col in downtime_df.columns]

            # Map downtime columns to standard names
            dt_col_map = {}
            for std_col, patterns in downtime_columns.items():
                for i, col_lower in enumerate(normalized_dt_columns):
                    orig_col = orig_dt_columns[i]

                    # Check for exact matches
                    if col_lower in [p.lower() for p in patterns]:
                        dt_col_map[orig_col] = std_col
                        break

                    # Check for partial matches
                    for pattern in patterns:
                        if pattern.lower() in col_lower:
                            dt_col_map[orig_col] = std_col
                            break

                    # Break if mapped this column
                    if orig_col in dt_col_map:
                        break

            # Apply downtime column mapping
            if dt_col_map:
                mapping_info = ", ".join(
                    [f"{orig} → {new}" for orig, new in dt_col_map.items()])
                # st.info(f"Downtime column mapping: {mapping_info}")
                downtime_df = downtime_df.rename(columns=dt_col_map)

        # Check for required columns in downtime data
        if downtime_df is None or downtime_df.empty \
                or 'date' not in downtime_df.columns \
                or 'downtime_pct' not in downtime_df.columns:
            st.error(
                "Downtime data must include non-empty 'date' and 'downtime_pct' columns.")
            return None, None, None

        # Convert dates to datetime
        batch_df['date'] = pd.to_datetime(batch_df['date'])
        downtime_df['date'] = pd.to_datetime(downtime_df['date'])

        # Standardize dates to end of month
        batch_df['date'] = batch_df['date'].apply(standardize_to_end_of_month)
        downtime_df['date'] = downtime_df['date'].apply(
            standardize_to_end_of_month)

        # Standardize downtime percentages to decimal format
        try:
            downtime_df['downtime_pct'] = downtime_df['downtime_pct'].apply(
                standardize_downtime_percentage)
        except ValueError as e:
            st.error(f"Error in downtime data: {str(e)}")
            return None, None, None

        # Set index for downtime DataFrame
        downtime_df.set_index('date', inplace=True)
        downtime_df.sort_index(inplace=True)

        # Get unique wells
        unique_wells = batch_df['well_name'].unique()
        if len(unique_wells) == 0:
            st.error("No wells found in batch forecast file.")
            return None, None, None

        # Create dictionary to store individual well data
        well_data = {}

        # Process each well
        for well in unique_wells:
            well_df = batch_df[batch_df['well_name'] == well][[
                'date', 'oil_rate', 'gas_rate', 'water_rate']]
            # Fill NaN values with zeros
            well_df['oil_rate'] = well_df['oil_rate'].fillna(0)
            well_df['gas_rate'] = well_df['gas_rate'].fillna(0)
            well_df['water_rate'] = well_df['water_rate'].fillna(0)
            well_df = well_df.set_index('date')
            well_df.sort_index(inplace=True)
            well_data[well] = well_df

        return well_data, downtime_df, unique_wells

    except Exception as e:
        st.error(f"Error processing batch data: {str(e)}")
        import traceback
        st.error(f"Traceback: {traceback.format_exc()}")
        return None, None, None


def validate_input_data(forecast_df, downtime_df):
    """Validate input dataframes."""
    errors = []
    warnings = []

    # Check forecast data
    possible_fluid_cols = ['oil_rate', 'gas_rate', 'water_rate']
    available_fluid_cols = [
        col for col in possible_fluid_cols if col in forecast_df.columns]

    # Check if at least one fluid type is present
    if not available_fluid_cols:
        errors.append(
            "At least one fluid type (oil_rate, gas_rate, or water_rate) \
                is required in forecast data.")
    else:
        # Check if any available fluid has data (non-zero sum)
        has_fluid_data = False
        for col in available_fluid_cols:
            if forecast_df[col].sum() > 0:
                has_fluid_data = True
                break

        if not has_fluid_data:
            errors.append("At least one fluid type must have non-zero values")

        # Add warnings for missing fluid types
        missing_fluids = [
            col.replace(
                '_rate',
                '') for col in possible_fluid_cols if col not in available_fluid_cols]
        if missing_fluids:
            warnings.append(
                f"Note: No {', '.join(missing_fluids)} data provided."
                "This is acceptable but may affect some calculations.")

    # Check downtime data
    if 'downtime_pct' not in downtime_df.columns:
        errors.append("Downtime data must include a 'downtime_pct' column")
    else:
        # Check if downtime percentages are within valid range (0-100)
        invalid_downtimes = downtime_df[downtime_df['downtime_pct'] > 1.0]
        if not invalid_downtimes.empty:
            errors.append(
                "Downtime percentages must be between 0 and 100 \
                (or 0 and 1 if in decimal format)")

    # Check date alignment between forecast and downtime data
    forecast_dates = set(forecast_df.index)
    downtime_dates = set(downtime_df.index)

    if not forecast_dates.issubset(downtime_dates):
        missing_dates = forecast_dates - downtime_dates
        errors.append(
            f"Missing downtime data for dates: {sorted(missing_dates)}")

    if not downtime_dates.issubset(forecast_dates):
        extra_dates = downtime_dates - forecast_dates
        warnings.append(
            f"Downtime data includes dates not in forecast: {sorted(extra_dates)}")

    return errors, warnings


def process_production_data(df):
    dtype_dict = {
        'oil_rate': np.float64,
        'gas_rate': np.float64,
        'water_rate': np.float64,
        'downtime_pct': np.float64,
        'liquid_rate': np.float64,
        'daysInMonth': np.int64,
        'oil_volume': np.float64,
        'gas_volume': np.float64,
        'water_volume': np.float64,
        'liquid_volume': np.float64,
        'Np': np.float64,
        'Gp': np.float64,
        'Wp': np.float64,
        'Lp': np.float64,
        'adj_oil_rate': np.float64,
        'adj_gas_rate': np.float64,
        'adj_water_rate': np.float64,
        'adj_liquid_rate': np.float64,
        'Np_out': np.float64,
        'Gp_out': np.float64,
        'Wp_out': np.float64,
        'Lp_out': np.float64
    }

    # Create additional columns
    df['liquid_rate'] = df['oil_rate'] + df['water_rate']
    df['GOR'] = np.where(
        df['gas_rate'] /
        df['oil_rate'] > 0,
        df['gas_rate'] /
        df['oil_rate'],
        0)
    df['WOR'] = np.where(
        df['water_rate'] /
        df['oil_rate'] > 0,
        df['water_rate'] /
        df['oil_rate'],
        0)

    # Calculate days in month based on the date format
    def get_days_in_month(date):
        if date.day == 1:  # Beginning of month
            return pd.Timestamp(date.year, date.month, 1).days_in_month
        else:  # End of month
            return date.days_in_month

    df['daysInMonth'] = df.index.map(get_days_in_month)

    df['oil_volume'] = (df['oil_rate'] * df['daysInMonth']) / 1000
    df['gas_volume'] = (df['gas_rate'] * df['daysInMonth']) / 1000
    df['water_volume'] = (df['water_rate'] * df['daysInMonth']) / 1000
    df['liquid_volume'] = (df['liquid_rate'] * df['daysInMonth']) / 1000

    df['Np'] = (df['oil_volume']).cumsum()
    df['Gp'] = (df['gas_volume']).cumsum()
    df['Wp'] = (df['water_volume']).cumsum()
    df['Lp'] = df['Np'] + df['Wp']

    df = df.replace([np.inf, -np.inf], np.nan).dropna(how='any')

    # Identify decline start
    if (df['oil_rate'] > 0).any():
        oil_rate_diff = np.diff(df['oil_rate'].values)
        decline_start_index = np.where(oil_rate_diff <= 0)[
            0][0] + 1 if len(np.where(oil_rate_diff <= 0)[0]) > 0 else 0
    elif (df['gas_rate'] > 0).any():
        gas_rate_diff = np.diff(df['gas_rate'].values)
        decline_start_index = np.where(gas_rate_diff <= 0)[
            0][0] + 1 if len(np.where(gas_rate_diff <= 0)[0]) > 0 else 0
    else:
        decline_start_index = 0

    df_interpolate = df.iloc[decline_start_index:][[
        'Np', 'oil_rate', 'GOR', 'WOR']]

    # Create output DataFrame with correct dtypes
    df_out = df.copy()

    # Initialize new columns with correct dtypes
    new_cols = [
        'adj_oil_rate',
        'adj_gas_rate',
        'adj_water_rate',
        'adj_liquid_rate',
        'adj_oil_volume',
        'adj_gas_volume',
        'adj_water_volume',
        'adj_liquid_volume',
        'Np_out',
        'Gp_out',
        'Wp_out',
        'Lp_out']
    for col in new_cols:
        df_out[col] = pd.Series(0.0, index=df_out.index, dtype='float64')

    df_out = df_out.astype(dtype_dict)

    # Perform the initial row calculations and assign them to df_out
    df_out.loc[df_out.index[0], 'adj_oil_rate'] = df_out.loc[df_out.index[0],
                                                             'oil_rate'] * (1 - df_out.loc[df_out.index[0], 'downtime_pct'])
    df_out.loc[df_out.index[0],
               'adj_gas_rate'] = df_out.loc[df_out.index[0],
                                            'GOR'] * df_out.loc[df_out.index[0],
                                                                'adj_oil_rate']
    df_out.loc[df_out.index[0],
               'adj_water_rate'] = df_out.loc[df_out.index[0],
                                              'WOR'] * df_out.loc[df_out.index[0],
                                                                  'adj_oil_rate']
    df_out.loc[df_out.index[0],
               'adj_liquid_rate'] = df_out.loc[df_out.index[0],
                                               'adj_oil_rate'] + df_out.loc[df_out.index[0],
                                                                            'adj_water_rate']
    df_out.loc[df_out.index[0],
               'adj_oil_volume'] = df_out.loc[df_out.index[0],
                                              'adj_oil_rate'] * df_out.loc[df_out.index[0],
                                                                           'daysInMonth'] / 1000
    df_out.loc[df_out.index[0], 'Np_out'] = (
        df_out.loc[df_out.index[0], 'adj_oil_volume']).cumsum()
    df_out.loc[df_out.index[0], 'Gp_out'] = (
        df_out.loc[df_out.index[0], 'adj_gas_volume']).cumsum()
    df_out.loc[df_out.index[0], 'Wp_out'] = (
        df_out.loc[df_out.index[0], 'adj_water_volume']).cumsum()
    df_out.loc[df_out.index[0],
               'Lp_out'] = df_out.loc[df_out.index[0],
                                      'Np_out'] + df_out.loc[df_out.index[0],
                                                             'Wp_out']

    # Initial conditions for the loop
    max_np = df['Np'].max()
    max_gp = df['Gp'].max()
    max_wp = df['Wp'].max()
    df['Lp'].max()

    last_np_out = 0
    last_gp_out = 0
    last_wp_out = 0
    current_date = df_out.index[0]

    # Main processing loop
    while True:
        if current_date not in df_out.index:
            new_row = {'oil_rate': 0,
                       'gas_rate': 0,
                       'water_rate': 0,
                       'downtime_pct': df_out['downtime_pct'].iloc[-1],
                       # Assume downtime_pct remains constant in extended years
                       'liquid_rate': 0,
                       'daysInMonth': get_days_in_month(current_date),
                       'oil_volume': 0,
                       'gas_volume': 0,
                       'water_volume': 0,
                       'liquid_volume': 0,
                       'Np': df_out['Np'].iloc[-1],
                       'Gp': df_out['Gp'].iloc[-1],
                       'Wp': df_out['Wp'].iloc[-1],
                       'Lp': df_out['Lp'].iloc[-1],
                       'adj_oil_rate': 0,
                       'adj_gas_rate': 0,
                       'adj_water_rate': 0,
                       'adj_liquid_rate': 0,
                       'adj_oil_volume': 0,
                       'adj_gas_volume': 0,
                       'adj_water_volume': 0,
                       'adj_liquid_volume': 0,
                       'Np_out': 0,
                       'Gp_out': 0,
                       'Wp_out': 0,
                       'Lp_out': 0
                       }

            # Date column is added to the dataframe and it should be the index
            df_out = pd.concat(
                [df_out, pd.DataFrame([new_row], index=[current_date])])
            df_out.reset_index(drop=True)

        # Check if we're still within the input range for non-declining oil
        # rates
        if current_date <= df.index[decline_start_index]:
            adj_oil_rate = df_out.loc[current_date, 'oil_rate'] * \
                (1 - df_out.loc[current_date, 'downtime_pct'])
            # Calculate gas and water rates based on input
            if adj_oil_rate > 0:
                adj_gas_rate = df_out.loc[current_date, 'gas_rate'] * \
                    (1 - df_out.loc[current_date, 'downtime_pct'])
                adj_water_rate = df_out.loc[current_date, 'water_rate'] * \
                    (1 - df_out.loc[current_date, 'downtime_pct'])
            else:
                adj_gas_rate = 0
                adj_water_rate = 0
            adj_liquid_rate = adj_oil_rate + adj_water_rate
        else:
            # Check if this date has all zero rates in the original dataframe
            if current_date in df.index and df.loc[current_date,
                                                   'oil_rate'] == 0 and df.loc[current_date,
                                                                               'gas_rate'] == 0 and df.loc[current_date,
                                                                                                           'water_rate'] == 0:
                # Honor zero rates only when all rates are zero
                adj_oil_rate = 0
                adj_gas_rate = 0
                adj_water_rate = 0
                adj_liquid_rate = 0
            else:
                # Use interpolated values for oil rate and adjust for downtime
                adj_oil_rate = interpolate_values(
                    last_np_out, df_interpolate['Np'], df_interpolate['oil_rate'])
                downtime_pct = df_out.loc[current_date, 'downtime_pct']
                adj_oil_rate *= (1 - downtime_pct)

                # Calculate gas and water rates
                adj_gas_rate = interpolate_values(
                    last_np_out, df_interpolate['Np'], df_interpolate['GOR']) * adj_oil_rate
                adj_water_rate = interpolate_values(
                    last_np_out, df_interpolate['Np'], df_interpolate['WOR']) * adj_oil_rate
                adj_liquid_rate = adj_oil_rate + adj_water_rate

        # No need to check for adj_liquid_rate existence anymore
        # if 'adj_liquid_rate' not in locals():
        #     adj_liquid_rate = adj_oil_rate + adj_water_rate

        # Update cumulative productions
        days_in_month = get_days_in_month(current_date)
        adj_oil_volume = adj_oil_rate * days_in_month / 1000
        adj_gas_volume = adj_gas_rate * days_in_month / 1000
        adj_water_volume = adj_water_rate * days_in_month / 1000
        adj_liquid_volume = adj_liquid_rate * days_in_month / 1000
        np_out = last_np_out + adj_oil_volume
        gp_out = last_gp_out + adj_gas_volume
        wp_out = last_wp_out + adj_water_volume
        lp_out = np_out + wp_out

        # Update existing row
        df_out.loc[current_date, 'adj_oil_rate'] = adj_oil_rate
        df_out.loc[current_date, 'adj_gas_rate'] = adj_gas_rate
        df_out.loc[current_date, 'adj_water_rate'] = adj_water_rate
        df_out.loc[current_date, 'adj_liquid_rate'] = adj_liquid_rate
        if adj_oil_rate != 0 and not np.isnan(adj_oil_rate):
            df_out.loc[current_date, 'WOR'] = adj_water_rate / \
                adj_oil_rate if adj_water_rate != 0 else 0
            df_out.loc[current_date, 'GOR'] = adj_gas_rate / \
                adj_oil_rate if adj_gas_rate != 0 else 0
        else:
            df_out.loc[current_date, 'WOR'] = 0
            df_out.loc[current_date, 'GOR'] = 0
        df_out.loc[current_date, 'adj_oil_volume'] = adj_oil_volume
        df_out.loc[current_date, 'adj_gas_volume'] = adj_gas_volume
        df_out.loc[current_date, 'adj_water_volume'] = adj_water_volume
        df_out.loc[current_date, 'adj_liquid_volume'] = adj_liquid_volume
        df_out.loc[current_date, 'Np_out'] = np_out
        df_out.loc[current_date, 'Gp_out'] = gp_out
        df_out.loc[current_date, 'Wp_out'] = wp_out
        df_out.loc[current_date, 'Lp_out'] = lp_out

        np_out = df_out.loc[current_date, 'Np_out']
        last_gp_out = df_out.loc[current_date, 'Gp_out']
        last_wp_out = df_out.loc[current_date, 'Wp_out']
        last_np_out = np_out

        # Move to next month based on current date format
        if current_date.day == 1:  # Beginning of month
            current_date = pd.Timestamp(
                current_date.year,
                current_date.month,
                1) + pd.offsets.MonthEnd(1)
        else:  # End of month
            current_date += pd.offsets.MonthEnd(1)

        # Check terminal conditions
        if last_np_out > max_np or last_gp_out > max_gp or last_wp_out > max_wp or len(
                df_out) >= 361:
            last_date = df_out.index[-2]
            current_date = df_out.index[-1]
            days_in_month = get_days_in_month(current_date)

            if last_np_out > max_np and last_gp_out > max_gp:
                # Both oil and gas exceeded - calculate rates to hit both
                # targets
                rem_np = last_np_out - max_np
                rem_gp = last_gp_out - max_gp

                # Calculate required rates to hit targets
                adj_oil_rate = rem_np * 1000 / days_in_month
                adj_gas_rate = rem_gp * 1000 / days_in_month
                last_wor = df_out.loc[last_date, 'WOR']
                adj_water_rate = adj_oil_rate * last_wor

            elif last_np_out > max_np:
                # Only oil exceeded
                rem_np = last_np_out - max_np
                adj_oil_rate = rem_np * 1000 / days_in_month
                last_gor = df_out.loc[last_date, 'GOR']
                adj_gas_rate = adj_oil_rate * last_gor
                last_wor = df_out.loc[last_date, 'WOR']
                adj_water_rate = adj_oil_rate * last_wor

            elif last_wp_out > max_wp:
                # Only water exceeded
                rem_wp = last_wp_out - max_wp
                adj_water_rate = rem_wp * 1000 / days_in_month
                adj_oil_rate = df_out.loc[last_date, 'adj_oil_rate']
                last_wor = df_out.loc[last_date, 'WOR']
                adj_gas_rate = adj_oil_rate * last_wor

            else:
                # Only gas exceeded
                rem_gp = last_gp_out - max_gp
                adj_gas_rate = rem_gp * 1000 / days_in_month
                adj_oil_rate = df_out.loc[last_date, 'adj_oil_rate']
                last_wor = df_out.loc[last_date, 'WOR']
                adj_water_rate = adj_oil_rate * last_wor

            # Update the rates in df_out
            df_out.loc[current_date, 'adj_oil_rate'] = adj_oil_rate
            df_out.loc[current_date, 'adj_gas_rate'] = adj_gas_rate
            df_out.loc[current_date, 'adj_water_rate'] = adj_water_rate
            df_out.loc[current_date,
                       'adj_liquid_rate'] = adj_oil_rate + adj_water_rate
            df_out.loc[current_date, 'WOR'] = adj_water_rate / \
                adj_oil_rate if adj_oil_rate > 0 else 0
            df_out.loc[current_date, 'GOR'] = adj_gas_rate / \
                adj_oil_rate if adj_oil_rate > 0 else 0

            # Calculate volumes
            adj_oil_volume = adj_oil_rate * days_in_month / 1000
            adj_gas_volume = adj_gas_rate * days_in_month / 1000
            adj_water_volume = adj_water_rate * days_in_month / 1000

            df_out.loc[current_date, 'adj_oil_volume'] = adj_oil_volume
            df_out.loc[current_date, 'adj_gas_volume'] = adj_gas_volume
            df_out.loc[current_date, 'adj_water_volume'] = adj_water_volume
            df_out.loc[current_date, 'adj_liquid_volume'] = (
                adj_oil_rate + adj_water_rate) * days_in_month / 1000

            # Set final cumulative volumes to exactly match targets
            df_out.loc[current_date,
                       'Np_out'] = max_np if last_np_out > max_np else last_np_out + adj_oil_volume
            df_out.loc[current_date,
                       'Gp_out'] = max_gp if last_gp_out > max_gp else last_gp_out + adj_gas_volume
            df_out.loc[current_date, 'Wp_out'] = last_wp_out + adj_water_volume
            df_out.loc[current_date,
                       'Lp_out'] = df_out.loc[current_date,
                                              'Np_out'] + df_out.loc[current_date,
                                                                     'Wp_out']

            # Add final timestep with zero rates but maintaining cumulative
            # volumes
            new_row = df_out.loc[current_date].copy()
            new_row[['adj_oil_rate', 'adj_gas_rate',
                     'adj_water_rate', 'adj_liquid_rate']] = 0
            new_row[['Np_out', 'Gp_out', 'Wp_out', 'Lp_out']] = [
                max_np, max_gp, new_row['Wp_out'], max_np + new_row['Wp_out']]
            new_df = pd.DataFrame(new_row).T.set_index(
                pd.Index([add_one_month(current_date)]))

            df_out = pd.concat([df_out, new_df])
            df_out.set_index(pd.Index(df_out.index), inplace=True)

        # Break condition
        if last_np_out > df_out['Np'].max(
        ) or last_gp_out > df_out['Gp'].max() or len(df_out) >= 361:
            break

        df_out['GOR_out'] = df_out['adj_gas_rate'] / \
            df_out['adj_oil_rate'].replace(0, np.nan)
        df_out['WOR_out'] = df_out['adj_water_rate'] / \
            df_out['adj_oil_rate'].replace(0, np.nan)

    return df_out, df_interpolate


def create_plotly_figure(df_out, well_name=None):
    """Create complete plotly figure with all traces."""
    if df_out is None or df_out.empty:
        # Return empty figure
        return go.Figure()
    if not well_name:
        well_name = ''
    title_text = f"Forecast Comparison ({well_name})" if well_name else "Forecast Comparison"
    fig = make_subplots(
        rows=4, cols=2,
        subplot_titles=(
            'Oil Rate vs Time', 'Cumulative Oil Production',
            'Gas Rate vs Time', 'Cumulative Gas Production',
            'Water Rate vs Time', 'Cumulative Water Production',
            'Liquid Rate vs Time', 'Cumulative Liquid Production'
        ),
        vertical_spacing=0.05,
        horizontal_spacing=0.1
    )

    # Common trace settings
    line_width = 3
    opacity = 0.8
    hover_tmpl = "%{y:,.0f}"

    # Calculate max values for axis scaling
    max_rate = max(
        df_out['oil_rate'].max(),
        df_out['adj_oil_rate'].max(),
        df_out['gas_rate'].max(),
        df_out['adj_gas_rate'].max(),
        df_out['water_rate'].max(),
        df_out['adj_water_rate'].max(),
        df_out['liquid_rate'].max(),
        df_out['adj_liquid_rate'].max()
    )

    max_cumulative = max(
        df_out['Np'].max(),
        df_out['Np_out'].max(),
        df_out['Gp'].max(),
        df_out['Gp_out'].max(),
        df_out['Wp'].max(),
        df_out['Wp_out'].max(),
        df_out['Lp'].max(),
        df_out['Lp_out'].max()
    )

    max_rate_rounded = max_rate * 1.1
    max_cumulative_rounded = max_cumulative * 1.1

    # Oil Rate and Cumulative
    fig.add_trace(
        go.Scatter(x=df_out.index, y=df_out['oil_rate'],
                   name='Oil Rate',
                   line=dict(color=COLORS['oil'], width=line_width),
                   opacity=opacity,
                   hovertemplate=hover_tmpl),
        row=1, col=1
    )
    fig.add_trace(
        go.Scatter(
            x=df_out.index,
            y=df_out['adj_oil_rate'],
            name='Oil Rate DT',
            line=dict(
                color=COLORS['oil'],
                width=line_width,
                dash='dot'),
            opacity=opacity,
            hovertemplate=hover_tmpl),
        row=1,
        col=1)
    fig.add_trace(
        go.Scatter(x=df_out.index, y=df_out['Np'],
                   name='Np',
                   line=dict(color=COLORS['oil'], width=line_width),
                   opacity=opacity,
                   hovertemplate=hover_tmpl),
        row=1, col=2
    )
    fig.add_trace(
        go.Scatter(
            x=df_out.index,
            y=df_out['Np_out'],
            name='Np DT',
            line=dict(
                color=COLORS['oil'],
                width=line_width,
                dash='dot'),
            opacity=opacity,
            hovertemplate=hover_tmpl),
        row=1,
        col=2)

    # Gas Rate and Cumulative
    fig.add_trace(
        go.Scatter(x=df_out.index, y=df_out['gas_rate'],
                   name='Gas Rate',
                   line=dict(color=COLORS['gas'], width=line_width),
                   opacity=opacity,
                   hovertemplate=hover_tmpl),
        row=2, col=1
    )
    fig.add_trace(
        go.Scatter(
            x=df_out.index,
            y=df_out['adj_gas_rate'],
            name='Gas Rate DT',
            line=dict(
                color=COLORS['gas'],
                width=line_width,
                dash='dot'),
            opacity=opacity,
            hovertemplate=hover_tmpl),
        row=2,
        col=1)
    fig.add_trace(
        go.Scatter(x=df_out.index, y=df_out['Gp'],
                   name='Gp',
                   line=dict(color=COLORS['gas'], width=line_width),
                   opacity=opacity,
                   hovertemplate=hover_tmpl),
        row=2, col=2
    )
    fig.add_trace(
        go.Scatter(
            x=df_out.index,
            y=df_out['Gp_out'],
            name='Gp DT',
            line=dict(
                color=COLORS['gas'],
                width=line_width,
                dash='dot'),
            opacity=opacity,
            hovertemplate=hover_tmpl),
        row=2,
        col=2)

    # Water Rate and Cumulative
    fig.add_trace(
        go.Scatter(x=df_out.index, y=df_out['water_rate'],
                   name='Water Rate',
                   line=dict(color=COLORS['water'], width=line_width),
                   opacity=opacity,
                   hovertemplate=hover_tmpl),
        row=3, col=1
    )
    fig.add_trace(
        go.Scatter(
            x=df_out.index,
            y=df_out['adj_water_rate'],
            name='Water Rate DT',
            line=dict(
                color=COLORS['water'],
                width=line_width,
                dash='dot'),
            opacity=opacity,
            hovertemplate=hover_tmpl),
        row=3,
        col=1)
    fig.add_trace(
        go.Scatter(x=df_out.index, y=df_out['Wp'],
                   name='Wp',
                   line=dict(color=COLORS['water'], width=line_width),
                   opacity=opacity,
                   hovertemplate=hover_tmpl),
        row=3, col=2
    )
    fig.add_trace(
        go.Scatter(
            x=df_out.index,
            y=df_out['Wp_out'],
            name='Wp DT',
            line=dict(
                color=COLORS['water'],
                width=line_width,
                dash='dot'),
            opacity=opacity,
            hovertemplate=hover_tmpl),
        row=3,
        col=2)

    # Liquid Rate and Cumulative
    fig.add_trace(
        go.Scatter(x=df_out.index, y=df_out['liquid_rate'],
                   name='Liquid Rate',
                   line=dict(color=COLORS['liquid'], width=line_width),
                   opacity=opacity,
                   hovertemplate=hover_tmpl),
        row=4, col=1
    )
    fig.add_trace(
        go.Scatter(
            x=df_out.index,
            y=df_out['adj_liquid_rate'],
            name='Liquid Rate DT',
            line=dict(
                color=COLORS['liquid'],
                width=line_width,
                dash='dot'),
            opacity=opacity,
            hovertemplate=hover_tmpl),
        row=4,
        col=1)
    fig.add_trace(
        go.Scatter(x=df_out.index, y=df_out['Lp'],
                   name='Lp',
                   line=dict(color=COLORS['liquid'], width=line_width),
                   opacity=opacity,
                   hovertemplate=hover_tmpl),
        row=4, col=2
    )
    fig.add_trace(
        go.Scatter(
            x=df_out.index,
            y=df_out['Lp_out'],
            name='Lp DT',
            line=dict(
                color=COLORS['liquid'],
                width=line_width,
                dash='dot'),
            opacity=opacity,
            hovertemplate=hover_tmpl),
        row=4,
        col=2)

    # Update layout and axes
    fig.update_layout(
        height=1400,
        width=1000,
        showlegend=False,
        title={
            'text': title_text,
            'y': 0.99,
            'x': 0.5,
            'xanchor': 'center',
            'yanchor': 'top',
            'font': {'size': 24}
        },
        **PLOT_LAYOUT
    )

    # Update axes with consistent ranges
    for i in range(1, 5):
        for j in range(1, 3):
            fig.update_xaxes(
                row=i, col=j,
                showgrid=True,
                gridwidth=1,
                gridcolor='rgba(128,128,128,0.2)',
                zeroline=False,
                showline=True,
                linewidth=1,
                linecolor='rgba(128,128,128,0.5)'
            )

            fig.update_yaxes(
                row=i,
                col=j,
                showgrid=True,
                gridwidth=1,
                gridcolor='rgba(128,128,128,0.2)',
                zeroline=False,
                showline=True,
                linewidth=1,
                linecolor='rgba(128,128,128,0.5)',
                tickformat=",",
                title_standoff=15,
                range=[
                    0,
                    max_rate_rounded if j == 1 else max_cumulative_rounded])

    return fig


def copy_to_clipboard(data_type='rates', all_wells=False):
    """Export data to clipboard or for download

    Args:
        data_type: 'rates' or 'volumes'
        all_wells: If True, export data for all processed wells, otherwise just the current well
    """
    if all_wells and st.session_state.processed_wells:
        # Create a buffer for Excel file
        buffer = io.BytesIO()

        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            # For each well, create sheets for rates and volumes
            for well_name, well_data in st.session_state.processed_wells.items():
                df = well_data['data']

                # Rates sheet
                rate_df = df[['adj_oil_rate', 'adj_gas_rate',
                              'adj_water_rate', 'adj_liquid_rate']].copy()
                rate_df.columns = [
                    'Oil Rate',
                    'Gas Rate',
                    'Water Rate',
                    'Liquid Rate']

                # Volumes sheet
                volume_df = df[['adj_oil_volume', 'adj_gas_volume',
                                'adj_water_volume', 'adj_liquid_volume']].copy()
                volume_df.columns = [
                    'Oil Volume',
                    'Gas Volume',
                    'Water Volume',
                    'Liquid Volume']

                # Write to Excel
                rate_df.to_excel(writer, sheet_name=f"{well_name[:15]}_Rates")
                volume_df.to_excel(
                    writer, sheet_name=f"{well_name[:15]}_Volumes")

        # Create a download button
        st.download_button(
            label="Download All Wells Excel",
            data=buffer.getvalue(),
            file_name=f"all_wells_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    elif 'processed_data' in st.session_state and st.session_state.processed_data is not None:
        df = st.session_state.processed_data

    if data_type == 'rates':
        temp_df = df[['adj_oil_rate', 'adj_gas_rate',
                      'adj_water_rate', 'adj_liquid_rate']].copy()
        temp_df.columns = ['Oil Rate', 'Gas Rate', 'Water Rate', 'Liquid Rate']
    else:
        temp_df = df[['adj_oil_volume', 'adj_gas_volume',
                      'adj_water_volume', 'adj_liquid_volume']].copy()
        temp_df.columns = [
            'Oil Volume',
            'Gas Volume',
            'Water Volume',
            'Liquid Volume']

        # Create CSV for download
        csv = temp_df.to_csv(index=True)

        # Create a download button
        st.download_button(
            label=f"Download {data_type.capitalize()} as CSV",
            data=csv,
            file_name=f"{data_type}_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.csv",
            mime="text/csv")

        # Display data in a text area for manual copying
        st.write(f"**{data_type.capitalize()} data (select all and copy):**")

        # Convert to tab-separated format for better clipboard pasting
        tsv_data = temp_df.to_csv(sep='\t', index=True)
        st.text_area(
            "Copyable TSV Data",
            tsv_data,
            height=150,
            label_visibility="collapsed")

        st.info(
            "👆 Select all text in the box above (Ctrl+A or Cmd+A), then copy (Ctrl+C or Cmd+C) to clipboard.")
    if 'processed_data' not in st.session_state or st.session_state.processed_data is None:
        st.error("No data available. Please process data first.")


def create_mosaic_template(
        df_out: pd.DataFrame,
        entity_name: str,
        reserve_category: str) -> pd.DataFrame:
    """Create Mosaic loader template from production data."""
    if df_out is None or df_out.empty:
        raise ValueError("No data to process.")
    # Get volume data (multiply by 1000 to convert back to bbl/mcf)
    vol_data = []

    # Process Oil volumes
    oil_data = df_out['adj_oil_volume'].copy() * 1000
    for date, value in oil_data.items():
        vol_data.append({
            'Entity Name': entity_name,
            'Reserve Category': reserve_category,
            'Use': 'Produced',
            'Product': 'Oil',
            'Detail Date (y-m-d)': date,
            'Cum For Period (Mcf or bbl)': value
        })

    # Process Gas volumes
    gas_data = df_out['adj_gas_volume'].copy() * 1000
    for date, value in gas_data.items():
        vol_data.append({
            'Entity Name': entity_name,
            'Reserve Category': reserve_category,
            'Use': 'Produced',
            'Product': 'Gas',
            'Detail Date (y-m-d)': date,
            'Cum For Period (Mcf or bbl)': value
        })

    # Process Water volumes
    water_data = df_out['adj_water_volume'].copy() * 1000
    for date, value in water_data.items():
        vol_data.append({
            'Entity Name': entity_name,
            'Reserve Category': reserve_category,
            'Use': 'Produced',
            'Product': 'Water',
            'Detail Date (y-m-d)': date,
            'Cum For Period (Mcf or bbl)': value
        })

    # Create DataFrame with exact column order
    mosaic_df = pd.DataFrame(vol_data)
    mosaic_df = mosaic_df[[
        'Entity Name',
        'Reserve Category',
        'Use',
        'Product',
        'Detail Date (y-m-d)',
        'Cum For Period (Mcf or bbl)'
    ]]

    product_order = pd.Categorical(
        mosaic_df['Product'], categories=[
            'Oil', 'Gas', 'Water'], ordered=True)
    mosaic_df['Product'] = product_order

    # Sort by Date and Product
    mosaic_df = mosaic_df.sort_values(
        by=['Product', 'Reserve Category', 'Detail Date (y-m-d)'], ascending=[True, True, True])

    mosaic_df['Detail Date (y-m-d)'] = mosaic_df['Detail Date (y-m-d)'].dt.date

    return mosaic_df


def generate_selected_wells_excel():
    import io
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Font
    from datetime import datetime
    selected_export_wells = st.session_state.get('export_wells', [])
    downtime_scenario = st.session_state.get('scenario_name', '')
    if not downtime_scenario:
        downtime_data = st.session_state.get('downtime_data', '')
        downtime_scenario = ''
        if downtime_data:
            try:
                downtime_df = pd.read_csv(io.StringIO(downtime_data), sep='\t')
                if len(downtime_df.columns) == 1:
                    downtime_df = pd.read_csv(
                        io.StringIO(downtime_data), sep=',')
                if not downtime_df.empty:
                    downtime_scenario = str(downtime_df.iloc[0, 0])
            except Exception:
                downtime_scenario = ''
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl', date_format='YYYY-MM-DD') as writer:
        for well in selected_export_wells:
            df = st.session_state.processed_wells[well]['data']
            date_col = pd.to_datetime(df.index).date
            export_df = pd.DataFrame({
                'Date': date_col,
                'Oil Rate (input)': df['oil_rate'],
                'Gas Rate (input)': df['gas_rate'],
                'Water Rate (input)': df['water_rate'],
                'Downtime Pct': df['downtime_pct'],
                'Oil Rate (DT)': df['adj_oil_rate'],
                'Gas Rate (DT)': df['adj_gas_rate'],
                'Water Rate (DT)': df['adj_water_rate'],
            })
            export_df.to_excel(
                writer,
                sheet_name=f"{well[:15]}",
                index=False,
                startrow=4)
    buffer.seek(0)
    wb = load_workbook(buffer)
    for well in selected_export_wells:
        ws = wb[f"{well[:15]}"]
        ws['A1'] = 'Date:'
        ws['B1'] = datetime.today().strftime('%Y-%m-%d')
        ws['A2'] = 'Well Name:'
        ws['B2'] = well
        ws['A3'] = 'Downtime Scenario:'
        ws['B3'] = downtime_scenario
        ws['A1'].font = Font(bold=True)
        ws['A2'].font = Font(bold=True)
        ws['A3'].font = Font(bold=True)
        for cell in ws[5]:
            cell.font = Font(bold=True)
        for row in ws.iter_rows(
                min_row=6,
                min_col=1,
                max_col=1,
                max_row=ws.max_row):
            for cell in row:
                cell.number_format = 'yyyy-mm-dd'
    out_buffer = io.BytesIO()
    wb.save(out_buffer)
    out_buffer.seek(0)
    filename = f"Forecast_downtime_adj_export_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return out_buffer.getvalue(), filename


# Custom CSS
st.markdown("""
    <style>
    .block-container {
        padding-top: 1rem;
        padding-bottom: 0.5rem;
        padding-left: 2.5rem !important;
        padding-right: 2.5rem !important;
    }
    .section-break {
        margin: 0.3rem 0;
    }
    .table-title {
        text-align: center;
        font-size: 20px;
        font-weight: bold;
        margin: 0.3rem;
    }
    .input-title {
        font-size: 20px;
        font-weight: 600;
        text-align: center;
        margin-bottom: 8px;
    }
    .plot-title {
        font-weight: bold;
        font-size: 24px;
        text-align: center;
        margin-bottom: 0.5rem;
    }
    .stButton > button, .stDownloadButton > button {
        font-weight: 700 !important;
        font-size: 10px !important;
        padding: 2px 8px !important;
        height: 28px !important;
        min-width: 0 !important;
        margin: 0 !important;
    }
    .tight-upload-row {
        margin-bottom: 0.1rem !important;
        margin-top: 0.1rem !important;
        padding: 0 !important;
    }
    .upload-instruction {
        font-size: 14px !important;
        margin-bottom: 0 !important;
        margin-top: 0 !important;
        padding: 0 !important;
        color: #444;
    }
    .big-center-btn {
        font-size: 18px !important;
        padding: 8px 24px !important;
        height: 42px !important;
        min-width: 120px !important;
        margin: 0 auto !important;
        display: block !important;
    }
    .sidebar-center {
        display: flex;
        flex-direction: column;
        align-items: center;
        gap: 0;
        margin-bottom: 0;
        margin-top: 0;
    }
    .stFileUploader {
        margin-top: 0 !important;
    }
    .well-dropdown {
        min-width: 150px !important;
        width: 100% !important;
    }
    </style>
""", unsafe_allow_html=True)

# Sidebar
with st.sidebar:
    st.header("Instructions")
    st.markdown("""
    1. Paste monthly forecast into tables or upload file
    2. Click 'Process Data' to run calculations
    3. Compare input and adjusted forecasts
    4. Export data as Excel file (optional)
    5. Generate Mosaic templates (optional)
    """)
    st.header("Data Format")
    st.markdown("""
        - Well Name (Entity Name if plan to use Mosaic templates)
        - Date, monthly
        - Oil Rate (bopd)
        - Gas Rate (mcfd)
        - Water Rate (bwpd)
        - Downtime Percentage (0-1 or 0-100)
    """)
    # Add template download buttons
    st.markdown('<div class="sidebar-center">', unsafe_allow_html=True)
    with open("template_production_forecast.xlsx", "rb") as f:
        st.download_button(
            label="Production Forecast Template",
            data=f.read(),
            file_name="template_production_forecast.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="sidebar_prod_template_btn")
    with open("template_downtime_forecast.xlsx", "rb") as f:
        st.download_button(
            label="Downtime Forecast Template",
            data=f.read(),
            file_name="template_downtime_forecast.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="sidebar_dt_template_btn")
    if st.button(
        "📚 View Documentation",
        use_container_width=True,
            key="sidebar_doc_btn"):
        st.session_state.show_docs = True
    st.markdown('</div>', unsafe_allow_html=True)

# Documentation Modal
if 'show_docs' not in st.session_state:
    st.session_state.show_docs = False

if st.session_state.show_docs:
    # Add top padding for the modal/expander
    st.markdown('<div style="height: 2.5rem"></div>', unsafe_allow_html=True)
    with st.expander("Documentation", expanded=True):
        # Add extra padding and larger font for the expander header
        st.markdown("""
            <style>
            .streamlit-expanderHeader {
                padding-top: 1.5rem !important;
                font-size: 1.5rem !important;
            }
            </style>
            """, unsafe_allow_html=True)
        st.markdown("## Process Flow & Interpolation Example")
        # Create two columns for the flowchart and plot
        flow_col, plot_col = st.columns([1, 1])

        with flow_col:
            # Process Flow diagram using Graphviz
            process_flow = """
            digraph {
                rankdir=TB;
                node [shape=box, style=filled, fillcolor=lightgray];

                Start [label="Start"];
                Input [label="Input Data"];
                Calc [label="Calculate Initial Rates"];
                Interp [label="Interpolate GOR/WOR vs Np"];
                Adjust [label="Adjust Rates for Downtime"];
                Check [label="Check Np"];
                End [label="End"];

                Start -> Input;
                Input -> Calc;
                Calc -> Interp;
                Interp -> Adjust;
                Adjust -> Check;
                Check -> Interp [label="Output Np < Input Np"];
                Check -> End [label="Target Reached"];
            }
            """
            st.graphviz_chart(process_flow)

        with plot_col:
            # Create data
            np_values = np.linspace(0, 20_000_000, 6)
            slope = 3.0 / 20_000_000
            intercept = -2.0
            log_wor_values = slope * np_values + intercept
            wor_values = np.exp(log_wor_values)

            np_interp = np.linspace(0, 20_000_000, 100)
            log_wor_interp = slope * np_interp + intercept
            wor_interp = np.exp(log_wor_interp)

            # Create figure with smaller size
            fig, ax = plt.subplots(figsize=(8, 5))

            # Plot data
            ax.plot(
                np_interp / 1e6,
                wor_interp,
                'r-',
                label='Forecast',
                linewidth=2)
            ax.plot(
                np_values / 1e6,
                wor_values,
                'bo',
                label='Interpolation Points',
                markersize=6)

            # Calculate the exact WOR value at Np = 12M
            wor_at_12m = np.exp(slope * 12_000_000 + intercept)

            # Add reference lines only up to the intersection point
            ax.plot([12, 12], [0.1, wor_at_12m], 'k--',
                    linewidth=1, alpha=0.5)  # Vertical line
            ax.plot([0, 12], [wor_at_12m, wor_at_12m], 'k--',
                    linewidth=1, alpha=0.5)  # Horizontal line

            # Add text
            ax.text(12.2, 0.15, 'WOR = 0.81', fontsize=10)

            # Set scales and limits
            ax.set_yscale('log')
            ax.set_ylim(0.1, 10)
            ax.set_xlim(0, 20)

            # Set grid
            ax.grid(True, which='both', linestyle='-', alpha=0.2)
            ax.grid(True, which='minor', linestyle=':', alpha=0.2)

            # Set labels
            ax.set_xlabel('Cumulative Oil Production (MMBO)')
            ax.set_ylabel('Water-Oil Ratio (WOR)')

            # Format y-axis ticks
            ax.yaxis.set_major_formatter(ticker.ScalarFormatter())
            ax.yaxis.set_minor_formatter(ticker.ScalarFormatter())

            # Add legend
            ax.legend()

            # Adjust layout
            plt.tight_layout()

            # Convert plot to image and display
            from io import BytesIO

            buf = BytesIO()
            plt.savefig(buf, format='png', dpi=300, bbox_inches='tight')
            buf.seek(0)
            st.image(buf)
            plt.close()

        # Close button for documentation
        if st.button("Close Documentation"):
            st.session_state.show_docs = False
            st.rerun()

# Main application
st.title("Forecast-Downtime Processing Tool")

# Two columns for input
col1, col2 = st.columns(2)

with col1:
    st.subheader("Production Forecast Input")
    well_name = st.text_input(
        "Well Name",
        placeholder="Enter well name...(leave blank if batch processing)",
        key="well_name_input",
        label_visibility="collapsed")
    if 'well_name' not in st.session_state:
        st.session_state.well_name = ""
    if well_name and well_name != st.session_state.well_name:
        st.session_state.well_name = well_name
        if 'forecast_data' in st.session_state and st.session_state.forecast_data:
            forecast_df = pd.read_csv(
                io.StringIO(
                    st.session_state.forecast_data),
                sep='\t')
            if 'Well Name' not in forecast_df.columns:
                forecast_df.insert(0, 'Well Name', well_name)
            else:
                forecast_df['Well Name'] = str(well_name)
            st.session_state.forecast_data = forecast_df.to_csv(
                sep='\t', index=False)
    forecast_columns = [
        'Well Name',
        'date',
        'oil_rate',
        'gas_rate',
        'water_rate']
    # Always use the latest session state for table display
    if 'forecast_data' in st.session_state and st.session_state.forecast_data:
        try:
            forecast_df = pd.read_csv(
                io.StringIO(
                    st.session_state.forecast_data),
                sep='\t')
            if len(forecast_df.columns) == 1:
                forecast_df = pd.read_csv(
                    io.StringIO(
                        st.session_state.forecast_data),
                    sep=',')
            for req in forecast_columns:
                if req not in forecast_df.columns:
                    forecast_df[req] = '' if req == 'Well Name' or req == 'date' else None
            forecast_df = forecast_df[forecast_columns]
        except Exception as e:
            st.error(f"Error parsing forecast data: {str(e)}")
            forecast_df = pd.DataFrame(columns=forecast_columns)
    else:
        forecast_df = pd.DataFrame({
            'Well Name': [well_name] * 6,
            'date': [''] * 6,
            'oil_rate': [None] * 6,
            'gas_rate': [None] * 6,
            'water_rate': [None] * 6
        })
    # --- Forecast Table ---
    # Ensure forecast_df has at least 6 rows for initial display
    if len(forecast_df) < 6:
        missing = 6 - len(forecast_df)
        add_df = pd.DataFrame({
            'Well Name': [well_name] * missing,
            'date': [''] * missing,
            'oil_rate': [None] * missing,
            'gas_rate': [None] * missing,
            'water_rate': [None] * missing
        })
        forecast_df = pd.concat([forecast_df, add_df], ignore_index=True)
    # Remove the maximum row limit since we want scrolling
    # but keep the initial rows setup

    # Keep the 6 rows display but allow scrolling
    edited_forecast_df = st.data_editor(
        forecast_df,
        num_rows="dynamic",  # Allow adding/removing rows
        use_container_width=True,
        hide_index=True,
        height=240,  # Fixed height for 6 rows
        column_config={
            "Well Name": st.column_config.TextColumn(
                "Well Name",
                help="Auto-filled from input above.",
                width="fixed",
                default=well_name,
            ),
            "date": st.column_config.TextColumn(
                "Date",
                help="Enter date in YYYY-MM-DD format",
                width="fixed",
                default="",
            ),
            "oil_rate": st.column_config.NumberColumn(
                "Oil Rate (bopd)",
                help="Enter oil rate in barrels per day (leave blank for 0)",
                min_value=0,
                format="%d",
                width="fixed",
                default=0,
            ),
            "gas_rate": st.column_config.NumberColumn(
                "Gas Rate (mcfd)",
                help="Enter gas rate in thousand cubic feet per day (leave blank for 0)",
                min_value=0,
                format="%d",
                width="fixed",
                default=0,
            ),
            "water_rate": st.column_config.NumberColumn(
                "Water Rate (bwpd)",
                help="Enter water rate in barrels per day (leave blank for 0)",
                min_value=0,
                format="%d",
                width="fixed",
                default=0,
            ),
        }
    )
    # Replace NaN values with zeros in the rate columns before saving to
    # session state
    if not edited_forecast_df.equals(forecast_df):
        edited_forecast_df['oil_rate'] = edited_forecast_df['oil_rate'].fillna(
            0)
        edited_forecast_df['gas_rate'] = edited_forecast_df['gas_rate'].fillna(
            0)
        edited_forecast_df['water_rate'] = edited_forecast_df['water_rate'].fillna(
            0)
        st.session_state.forecast_data = edited_forecast_df.to_csv(
            sep='\t', index=False)
        forecast_df = edited_forecast_df.copy()
    st.markdown(
        "<div class='upload-instruction'>Upload Excel file (.xlsx)</div>",
        unsafe_allow_html=True)
    uploaded_file = st.file_uploader(
        "Upload file",
        type=['xlsx'],
        key="forecast_uploader",
        label_visibility="collapsed")
    # --- NEW: Process file upload immediately ---
    if uploaded_file is not None:
        try:
            df = pd.read_excel(uploaded_file)

            if df is not None and not df.empty:
                # Create a mapping of original columns to standard names
                standard_columns = {
                    'Well Name': [
                        'well_name', 'well name', 'wellname', 'well', 'name', 'entity', 'Well Name'], 'date': [
                        'date', 'Date'], 'oil_rate': [
                        'oil_rate', 'oil rate', 'oilrate', 'oil', 'Oil Rate'], 'gas_rate': [
                        'gas_rate', 'gas rate', 'gasrate', 'gas', 'Gas Rate'], 'water_rate': [
                        'water_rate', 'water rate', 'waterrate', 'wtrrate', 'water', 'wtr', 'Water Rate']}

                # Preserve original column names for reporting
                orig_columns = list(df.columns)

                # Create a normalized version of columns for matching (strip,
                # lowercase)
                normalized_columns = [str(col).strip().lower()
                                      for col in df.columns]

                # Map columns to standard names
                col_map = {}
                for std_col, patterns in standard_columns.items():
                    for i, col_lower in enumerate(normalized_columns):
                        orig_col = orig_columns[i]

                        # First, check for exact match with any pattern
                        if col_lower in [p.lower() for p in patterns]:
                            col_map[orig_col] = std_col
                            break

                        # Then check for partial matches (skip if already
                        # mapped)
                        if orig_col not in col_map:
                            for pattern in patterns:
                                if pattern.lower() in col_lower:
                                    col_map[orig_col] = std_col
                                    break

                        # Break out of the loop if we've mapped this column
                        if orig_col in col_map:
                            break

                # Report mapping for debugging
                if col_map and DEBUG_MODE:
                    mapping_info = ", ".join(
                        [f"{orig} → {new}" for orig, new in col_map.items()])
                    st.info(f"Column mapping: {mapping_info}")

                # Apply column mapping
                df = df.rename(columns=col_map)

                # Add Well Name if missing and well_name is provided
                if 'Well Name' not in df.columns and 'well_name' in df.columns:
                    df['Well Name'] = df['well_name']
                    if DEBUG_MODE:
                        st.info("Mapped 'well_name' to 'Well Name'")
                elif 'Well Name' not in df.columns and well_name:
                    df.insert(0, 'Well Name', well_name)
                elif 'Well Name' in df.columns and well_name \
                        and (df['Well Name'].isnull().all() or (df['Well Name'] == '').all()):
                    df['Well Name'] = well_name

                required_cols = [
                    'Well Name',
                    'date',
                    'oil_rate',
                    'gas_rate',
                    'water_rate']
                missing_cols = [
                    col for col in required_cols if col not in df.columns]

                if missing_cols:
                    st.error(
                        f"Forecast file is missing required columns: {', '.join(missing_cols)}. Please check your file.")
                    if DEBUG_MODE:
                        st.info(f"Available columns: {', '.join(df.columns)}")
                elif df.empty:
                    st.error("Forecast file is empty. Please check your file.")
                else:
                    st.session_state.forecast_data = df.to_csv(
                        sep='\t', index=False)
                    # Refresh the table and show success message
                    st.success(
                        f"Successfully loaded {len(df['Well Name'].unique())} wells.")
                    # Fill NaN values with zeros before assigning
                    df['oil_rate'] = df['oil_rate'].fillna(0)
                    df['gas_rate'] = df['gas_rate'].fillna(0)
                    df['water_rate'] = df['water_rate'].fillna(0)
                    forecast_df = df[required_cols]
        except Exception as e:
            st.error(f"Error reading Excel file: {str(e)}")

with col2:
    st.subheader("Downtime Forecast Input")
    scenario_name = st.text_input(
        "Downtime Scenario",
        placeholder="Enter downtime scenario (i.e., OL3)...(leave blank if loading from file)",
        key="scenario_input",
        label_visibility="collapsed")
    if 'scenario_name' not in st.session_state:
        st.session_state.scenario_name = ""
    if scenario_name and scenario_name != st.session_state.scenario_name:
        st.session_state.scenario_name = scenario_name
        if 'downtime_data' in st.session_state and st.session_state.downtime_data:
            downtime_df = pd.read_csv(
                io.StringIO(
                    st.session_state.downtime_data),
                sep='\t')
            if 'Scenario' not in downtime_df.columns:
                downtime_df.insert(0, 'Scenario', scenario_name)
            else:
                downtime_df['Scenario'] = str(scenario_name)
            st.session_state.downtime_data = downtime_df.to_csv(
                sep='\t', index=False)
    downtime_columns = ['Scenario', 'date', 'downtime_pct']
    # Always use the latest session state for table display
    if 'downtime_data' in st.session_state and st.session_state.downtime_data:
        try:
            downtime_df = pd.read_csv(
                io.StringIO(
                    st.session_state.downtime_data),
                sep='\t')
            if len(downtime_df.columns) == 1:
                downtime_df = pd.read_csv(
                    io.StringIO(
                        st.session_state.downtime_data),
                    sep=',')
            for req in downtime_columns:
                if req not in downtime_df.columns:
                    downtime_df[req] = '' if req == 'Scenario' or req == 'date' else None
            downtime_df = downtime_df[downtime_columns]
            downtime_df['Scenario'] = downtime_df['Scenario'].astype(str)
        except Exception as e:
            st.error(f"Error parsing downtime data: {str(e)}")
            downtime_df = pd.DataFrame(columns=downtime_columns)
    else:
        downtime_df = pd.DataFrame({
            'Scenario': [scenario_name] * 6,
            'date': [''] * 6,
            'downtime_pct': [None] * 6
        })
    # --- Downtime Table ---
    # Keep the code to ensure downtime_df has at least 6 rows initially
    # --- Downtime Table ---
    # Ensure downtime_df has at least 6 rows for initial display
    if len(downtime_df) < 6:
        missing = 6 - len(downtime_df)
        add_df = pd.DataFrame({
            'Scenario': [scenario_name] * missing,
            'date': [''] * missing,
            'downtime_pct': [None] * missing
        })
        downtime_df = pd.concat([downtime_df, add_df], ignore_index=True)
    # Remove the maximum row limit since we want scrolling
    # but keep the initial rows setup

    # Always ensure downtime_df has exactly 6 visible rows but allow scrolling
    edited_downtime_df = st.data_editor(
        downtime_df,
        num_rows="dynamic",  # Change from "fixed" to "dynamic"
        use_container_width=True,
        hide_index=True,
        height=240,  # Fixed height for 6 rows
        column_config={
            "Scenario": st.column_config.TextColumn(
                "Scenario",
                help="Auto-filled from input above.",
                width="fixed",
                default=scenario_name,
            ),
            "date": st.column_config.TextColumn(
                "Date",
                help="Enter date in YYYY-MM-DD format",
                width="fixed",
                default="",
            ),
            "downtime_pct": st.column_config.NumberColumn(
                "Downtime Percentage",
                help="Enter downtime percentage (0-100 or 0-1, leave blank for 0)",
                min_value=0,
                max_value=100,
                format="%.2f",
                width="fixed",
                default=0,
            ),
        }
    )
    # Replace NaN values with zeros in the downtime_pct column before saving
    # to session state
    if not edited_downtime_df.equals(downtime_df):
        edited_downtime_df['downtime_pct'] = edited_downtime_df['downtime_pct'].fillna(
            0)
        st.session_state.downtime_data = edited_downtime_df.to_csv(
            sep='\t', index=False)
        downtime_df = edited_downtime_df.copy()
    st.markdown(
        "<div class='upload-instruction'>Upload Excel file (.xlsx)</div>",
        unsafe_allow_html=True)
    uploaded_downtime = st.file_uploader(
        "Upload file",
        type=['xlsx'],
        key="downtime_uploader",
        label_visibility="collapsed")
    # --- NEW: Process downtime file upload immediately ---
    if uploaded_downtime is not None:
        try:
            df = pd.read_excel(uploaded_downtime)

            if df is not None and not df.empty:
                # Create a mapping of original columns to standard names
                standard_columns = {
                    'Scenario': [
                        'scenario',
                        'scenario_name',
                        'case',
                        'name',
                        'downtime_scenario',
                        'Scenario'],
                    'date': [
                        'date',
                        'Date'],
                    'downtime_pct': [
                        'downtime_pct',
                        'downtime',
                        'dt_pct',
                        'dt',
                        'pct',
                        'percent',
                        'downtime_percent']}

                # Preserve original column names for reporting
                orig_columns = list(df.columns)

                # Create a normalized version of columns for matching (strip,
                # lowercase)
                normalized_columns = [str(col).strip().lower()
                                      for col in df.columns]

                # Map columns to standard names
                col_map = {}
                for std_col, patterns in standard_columns.items():
                    for i, col_lower in enumerate(normalized_columns):
                        orig_col = orig_columns[i]

                        # First, check for exact match with any pattern
                        if col_lower in [p.lower() for p in patterns]:
                            col_map[orig_col] = std_col
                            break

                        # Then check for partial matches (skip if already
                        # mapped)
                        if orig_col not in col_map:
                            for pattern in patterns:
                                if pattern.lower() in col_lower:
                                    col_map[orig_col] = std_col
                                    break

                        # Break out of the loop if we've mapped this column
                        if orig_col in col_map:
                            break

                # Report mapping for debugging
                if col_map and DEBUG_MODE:
                    mapping_info = ", ".join(
                        [f"{orig} → {new}" for orig, new in col_map.items()])
                    st.info(f"Downtime column mapping: {mapping_info}")

                # Apply column mapping
                df = df.rename(columns=col_map)

                # Add Scenario if missing
                if 'Scenario' not in df.columns and scenario_name:
                    df.insert(0, 'Scenario', scenario_name)

                required_cols = ['Scenario', 'date', 'downtime_pct']
                missing_cols = [
                    col for col in required_cols if col not in df.columns]

                if missing_cols:
                    st.error(
                        f"Downtime file is missing required columns: {', '.join(missing_cols)}. Please check your file.")
                    if DEBUG_MODE:
                        st.info(f"Available columns: {', '.join(df.columns)}")
                elif df.empty:
                    st.error("Downtime file is empty. Please check your file.")
                else:
                    st.session_state.downtime_data = df.to_csv(
                        sep='\t', index=False)
                    # Refresh the table and show success message
                    st.success("Successfully loaded downtime data.")
                    # Fill NaN values with zeros before assigning
                    df['downtime_pct'] = df['downtime_pct'].fillna(0)
                    downtime_df = df[required_cols]
        except Exception as e:
            st.error(f"Error reading Excel file: {str(e)}")

show_data_error = False
if not st.session_state.forecast_data or not st.session_state.downtime_data:
    show_data_error = True

colA, colB = st.columns(2)

if 'show_data_error' not in st.session_state:
    st.session_state.show_data_error = False


def handle_process_data():
    if not st.session_state.forecast_data or not st.session_state.downtime_data:
        st.session_state.show_data_error = True
    else:
        st.session_state.show_data_error = False
        process_and_store_data()


if st.session_state.show_data_error:
    st.error("Please provide both forecast and downtime data.")

with colA:
    st.button(
        "Process Data",
        type="primary",
        key="process_single",
        on_click=handle_process_data)
with colB:
    def clear_input():
        st.session_state.forecast_data = ""
        st.session_state.downtime_data = ""
        st.session_state.processed_data = None
        st.session_state.figure = None
        st.session_state.batch_data = None
        st.session_state.batch_wells = []
        st.session_state.selected_well = None
        st.session_state.well_forecast_data = {}
        st.session_state.processed_wells = {}
        st.session_state.well_name = ""
        st.session_state.scenario_name = ""
        st.session_state.well_name_input = ""
        st.session_state.scenario_input = ""
        st.session_state.show_data_error = False
    st.button("Clear Input", key="clear_single", on_click=clear_input)


def process_and_store_batch_data():
    """Process batch input data and store results in session state"""
    if not st.session_state.batch_data or not st.session_state.downtime_data:
        st.error("Please provide both batch forecast and downtime data.")
        return False

    try:
        # Process batch data
        well_data, downtime_df, unique_wells = process_batch_data(
            st.session_state.batch_data, st.session_state.downtime_data)
        if well_data is None or downtime_df is None:
            return False

        # Store the well names in session state
        st.session_state.batch_wells = list(unique_wells)

        # Store the well data for later use
        st.session_state.well_forecast_data = well_data

        # Set the first well as selected
        if not st.session_state.selected_well and len(unique_wells) > 0:
            st.session_state.selected_well = unique_wells[0]

        # Process each well
        st.info(f"Processing {len(unique_wells)} wells...")
        progress_bar = st.progress(0)

        for i, well_name in enumerate(unique_wells):
            # Update progress
            progress_bar.progress((i + 1) / len(unique_wells))

            # Get well forecast data
            well_forecast = well_data[well_name]

            # Validate well data
            errors, warnings = validate_input_data(well_forecast, downtime_df)
            if errors:
                st.warning(
                    f"Skipping {well_name} due to errors: {'; '.join(errors)}")
                continue

            # Combine and process data
            df = well_forecast.join(downtime_df, how='left')
            df = df.sort_index()
            try:
                df_out, df_interpolate = process_production_data(df)
                # Create figure for this well
                well_figure = create_plotly_figure(df_out, well_name=well_name)
                # Store in processed wells
                st.session_state.processed_wells[well_name] = {
                    'data': df_out,
                    'figure': well_figure
                }
                # If this is the selected well, update the main figure and data
                if well_name == st.session_state.selected_well:
                    st.session_state.figure = well_figure
                    st.session_state.processed_data = df_out
            except Exception as e:
                st.warning(f"Error processing {well_name}: {str(e)}")
                continue
        st.success(
            f"Successfully processed {len(st.session_state.processed_wells)} wells")
        return True

    except Exception as e:
        st.error(f"Error processing batch data: {str(e)}")
        return False


# Display results if data is processed
st.markdown("<div class='section-break'></div>", unsafe_allow_html=True)
if st.session_state.processed_data is not None and 'processed_wells' in st.session_state and st.session_state.processed_wells:
    df_out = st.session_state.processed_data
    st.markdown("## Results")
    # Well selection dropdown (always show if processed_wells exists)
    well_options = list(st.session_state.processed_wells.keys())
    if well_options:
        # 1.5:5.5 ratio (50% wider than original 1:6)
        dropdown_col, _ = st.columns([1.5, 5.5])
        with dropdown_col:
            # Apply CSS to the well dropdown
            st.markdown('''
                <style>
                [data-testid="stSelectbox"] > div:first-child > div:first-child {
                    min-width: 150px;
                }
                </style>
                ''', unsafe_allow_html=True)
            selected_well = st.selectbox(
                "Select Well to View",
                well_options,
                index=well_options.index(
                    st.session_state.selected_well) if st.session_state.selected_well in well_options else 0,
                key="well_dropdown")
        if selected_well != st.session_state.selected_well:
            st.session_state.selected_well = selected_well
            st.session_state.processed_data = st.session_state.processed_wells[
                selected_well]['data']
            st.session_state.figure = st.session_state.processed_wells[selected_well]['figure']
    else:
        selected_well = st.session_state.selected_well

    # Update the figure with the selected well name
    selected_well_name = st.session_state.selected_well if 'selected_well' in st.session_state else ''
    st.session_state.figure = create_plotly_figure(
        st.session_state.processed_data, well_name=selected_well_name)
    st.plotly_chart(st.session_state.figure, use_container_width=True)

    # --- Well-Year Summary Table ---
    selected_well_name = st.session_state.selected_well if 'selected_well' in st.session_state else ''
    summary_rows = []
    if selected_well_name and selected_well_name in st.session_state.processed_wells:
        well_data = st.session_state.processed_wells[selected_well_name]
        df = well_data['data'].copy()
        df['year'] = df.index.year
        # Input rates/volumes
        df['oil_volume_input'] = df['oil_rate'] * df['daysInMonth'] / 1000
        df['gas_volume_input'] = df['gas_rate'] * df['daysInMonth'] / 1000
        df['water_volume_input'] = df['water_rate'] * df['daysInMonth'] / 1000
        df['boe_input'] = df['oil_rate'] + \
            df['gas_rate'] / 6 + df['water_rate']
        df['boe_processed'] = df['adj_oil_rate'] + \
            df['adj_gas_rate'] / 6 + df['adj_water_rate']
        df['oil_volume_processed'] = df['adj_oil_rate'] * df['daysInMonth'] / 1000
        df['gas_volume_processed'] = df['adj_gas_rate'] * df['daysInMonth'] / 1000
        df['water_volume_processed'] = df['adj_water_rate'] * \
            df['daysInMonth'] / 1000
        for year, group in df.groupby('year'):
            summary_rows.append({
                'Year': int(year),
                'Avg Oil Rate': group['oil_rate'].mean(),
                'Avg Gas Rate': group['gas_rate'].mean(),
                'Avg Water Rate': group['water_rate'].mean(),
                'Avg BOE/d': group['boe_input'].mean(),
                'BOE Vol, MBOE': group['boe_input'].sum(),
                'Avg Downtime': group['downtime_pct'].mean() * 100,
                'Avg Oil Rate (DT)': group['adj_oil_rate'].mean(),
                'Avg Gas Rate (DT)': group['adj_gas_rate'].mean(),
                'Avg Water Rate (DT)': group['adj_water_rate'].mean(),
                'Avg BOE/d (DT)': group['boe_processed'].mean(),
                'BOE Vol, MBOE (DT)': group['boe_processed'].sum(),
            })
    summary_df = pd.DataFrame(summary_rows)
    display_cols = [
        'Year',
        'Avg Oil Rate',
        'Avg Gas Rate',
        'Avg Water Rate',
        'Avg BOE/d',
        'BOE Vol, MBOE',
        'Avg Downtime',
        'Avg Oil Rate (DT)',
        'Avg Gas Rate (DT)',
        'Avg Water Rate (DT)',
        'Avg BOE/d (DT)',
        'BOE Vol, MBOE (DT)']
    summary_df = summary_df[display_cols]
    # Format numbers
    for col in summary_df.columns:
        if summary_df[col].dtype in [float, np.float64]:
            if 'Downtime' in col:
                summary_df[col] = summary_df[col].map('{:.2f}%'.format)
            elif col == 'Year':
                summary_df[col] = summary_df[col].astype(int)
            else:
                summary_df[col] = summary_df[col].map('{:,.0f}'.format)
    # Remove any extra blank row in the summary table
    summary_df = summary_df[summary_df['Year'].notna()]
    summary_df = summary_df[summary_df['Year'].apply(
        lambda x: isinstance(x, (int, np.integer)))]
    st.markdown(
        f'<p class="table-title">Well-Year Summary Table ({selected_well_name})</p>',
        unsafe_allow_html=True)
    st.dataframe(
        summary_df,
        use_container_width=True,
        height=40 *
        len(summary_df) +
        40,
        hide_index=True)

    # --- Export Section ---
    st.markdown("<div style='margin: 1rem 0;'></div>", unsafe_allow_html=True)
    st.markdown("**Export Options**")
    all_wells = list(st.session_state.processed_wells.keys())
    selected_export_wells = st.multiselect(
        "Select wells to export (rates/volumes)",
        all_wells,
        default=all_wells,
        key="export_wells")

    # Generate and download in one go without using rerun
    export_col1, _ = st.columns([1, 1])
    with export_col1:
        # Generate the Excel file when button is clicked
        if st.button(
            "Generate and Download Excel",
            key="generate_export",
            help="Generate and download Excel file with selected wells",
                type="primary"):
            excel_bytes, excel_filename = generate_selected_wells_excel()

            # Create a download button for the generated file
            st.download_button(
                label="Click to Download Excel",
                data=excel_bytes,
                file_name=excel_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="download_excel")

    # --- Mosaic Export Section ---
    st.markdown("<div style='margin: 1rem 0;'></div>", unsafe_allow_html=True)
    st.markdown("**Mosaic Export Options**")
    selected_mosaic_wells = st.multiselect(
        "Select wells for Mosaic export",
        all_wells,
        default=all_wells,
        key="mosaic_export_wells")
    reserves_cycle = st.text_input(
        "Reserves Cycle Label (i.e., MY25, YE25)",
        placeholder="Enter reserves cycle label...(optional)",
        key="mosaic_reserves_cycle")
    reserve_category = st.selectbox(
        "Reserve Category", [
            "PDP", "2PDP", "3PDP"], key="mosaic_reserve_cat")

    # Generate and download in one go without using rerun
    mosaic_col1, _ = st.columns([1, 1])
    with mosaic_col1:
        # Generate the Mosaic templates when button is clicked
        if st.button(
            "Generate and Download Mosaic Templates",
            key="export_selected_mosaic",
            help="Generate and download Mosaic templates for selected wells",
                type="primary"):
            import io
            # import zipfile  # REMOVE this line to fix F811
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w') as zf:
                for well in selected_mosaic_wells:
                    mosaic_df = create_mosaic_template(
                        df_out=st.session_state.processed_wells[well]['data'],
                        entity_name=well,
                        reserve_category=reserve_category
                    )
                    excel_buffer = io.BytesIO()
                    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                        mosaic_df.to_excel(
                            writer, index=False, sheet_name='Sheet1')
                    excel_buffer.seek(0)
                    today_str = pd.Timestamp.now().strftime('%Y%m%d')
                    filename = f"{reserve_category}_{well}_Mosaic_Template_{today_str}.xlsx"
                    zf.writestr(filename, excel_buffer.getvalue())

            # Create the download button for the generated ZIP
            zip_buffer.seek(0)
            zip_filename = f"Mosaic_Templates_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.zip"

            st.download_button(
                label="Click to Download Mosaic Templates",
                data=zip_buffer.getvalue(),
                file_name=zip_filename,
                mime="application/zip",
                key="download_mosaic"
            )
